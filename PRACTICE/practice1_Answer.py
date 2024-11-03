'''
Author: gumpcpy gumpcpy@gmail.com
Date: 2024-11-03 13:42:31
LastEditors: gumpcpy gumpcpy@gmail.com
LastEditTime: 2024-11-03 16:23:30
Description: 
1-
把一張圖片放到excel裡面

2-
把圖片寬度固定在 350  圖片的長寬比維持原圖

3-
把一個指定folder裡面的png,jpg依照名稱順序 寫入excel B欄位裡
且A欄位為檔名

4-
一切和3一樣 但是含有圖片的路徑由使用者輸入
然後 excel 存到含有圖片的路徑下

'''
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as xlImage
from PIL import Image
import shutil


def step1():

    wb = Workbook()
    ws = wb.active
    ws.append(['Filename','Image'])

    img_path = '/Users/gump/Desktop/img/a.png'
    img = Image.open(img_path)
    print(img.height)

    ws.add_image(xlImage(img),'B2')

    wb.save('./Test.xlsx')
    print("DONE!")

def step2():
    wb = Workbook()
    ws = wb.active
    ws.append(['Filename', 'Image'])

    img_path = '/Users/gump/Desktop/img/a.png'
    img = Image.open(img_path)
    print(img.height)
    width_percentage = float(350 / img.width)
    resize_img_height = int(img.height * width_percentage)
    img = img.resize((350, resize_img_height), Image.LANCZOS)
    img = img.convert("RGB")
    img.save('./img.jpg')

    ws.add_image(xlImage('./img.jpg'), 'B2')

    wb.save('./Test.xlsx')

    os.remove('./img.jpg')
    print("DONE!")


def step3():
    wb = Workbook()
    ws = wb.active
    ws.append(['Filename', 'Image'])

    folder = '/Users/gump/Desktop/img'
    allowed_ext = ['.jpg','.png','.jpeg']

    # tmp_folder for tmp resized img
    tmp_folder = './tmp'
    if os.path.exists(tmp_folder):
        # 刪除已存在的 tmp 資料夾及其內容
        shutil.rmtree(tmp_folder)
    # 創建新的 tmp 資料夾
    os.makedirs(tmp_folder)

    #read img in folder to img_list and sort by name
    index = 1
    for file in os.listdir(folder):
        if file.lower().endswith(tuple(allowed_ext)):
            index += 1
            img = Image.open(os.path.join(folder, file))
            width_percentage = float(350 / img.width)
            resize_img_height = int(img.height * width_percentage)
            img = img.resize((350, resize_img_height), Image.LANCZOS)
            img = img.convert("RGB")
            img.save(f"{tmp_folder}/img_{index}.jpg")

            ws.cell(row=index, column=1, value=file)
            ws.add_image(xlImage(f"{tmp_folder}/img_{index}.jpg"), f"B{index}")
            ws.row_dimensions[index].height = img.height / 1.333
            

    wb.save('./Test.xlsx')
    
    shutil.rmtree(tmp_folder)
    

def step4(path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Filename', 'Image'])

    folder = path
    allowed_ext = ['.jpg', '.png', '.jpeg']

    # tmp_folder for tmp resized img
    tmp_folder = './tmp'
    if os.path.exists(tmp_folder):
        # 刪除已存在的 tmp 資料夾及其內容
        shutil.rmtree(tmp_folder)
    # 創建新的 tmp 資料夾
    os.makedirs(tmp_folder)

    # read img in folder to img_list and sort by name
    index = 1
    for file in os.listdir(folder):
        if file.lower().endswith(tuple(allowed_ext)):
            index += 1
            img = Image.open(os.path.join(folder, file))
            width_percentage = float(350 / img.width)
            resize_img_height = int(img.height * width_percentage)
            img = img.resize((350, resize_img_height), Image.LANCZOS)
            img = img.convert("RGB")
            img.save(f"{tmp_folder}/img_{index}.jpg")

            ws.cell(row=index, column=1, value=file)
            ws.add_image(xlImage(f"{tmp_folder}/img_{index}.jpg"), f"B{index}")
            ws.row_dimensions[index].height = img.height / 1.333

    wb.save(os.path.join(path,'Image.xlsx'))

    shutil.rmtree(tmp_folder)


if __name__ == '__main__':
    path = input("拖入含有圖片的文件夾:").strip()
    step4(path)

